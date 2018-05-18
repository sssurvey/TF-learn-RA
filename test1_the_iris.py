import openpyxl
import sys
from sklearn import tree

dataSet_iris_flower = openpyxl.load_workbook('test_data_sets/fishers_iris_data.xlsx')
sheet1 = dataSet_iris_flower['Sheet1']

def collectData(sheet1):
    """features = []
    labels = []
    clf = tree.DecisionTreeClassifier(features,labels)
    clf = clf.fit(features,labels)
    """
    dataCata = {}
    tempList = []
    for lines_loop in range(1,151):
        for columns_loop in range (2,7):
            tempList.append(str(sheet1.cell(row=lines_loop,column=columns_loop).value))
        dataCata[lines_loop] = tempList
        tempList = []
    # print(dataCata.values())
    return dataCata

def analysisData(dataCata):
    features = []
    labels = []
    for values in dataCata.values():
        features.append(values[0:4])
        labels.append(values[-1])
    """now both the labels and features are set"""
    clf = tree.DecisionTreeClassifier()
    clf = clf.fit(features,labels)
    return clf

def main():
    # test_out = sheet1.cell(row=1, column=6).value
    # print(test_out)
    temp_data = collectData(sheet1)
    temp_tree = analysisData(temp_data)
    print(temp_tree.predict([[3.2, 3.1, 1.6, 0.2]]))

if __name__ == "__main__":
    main()
